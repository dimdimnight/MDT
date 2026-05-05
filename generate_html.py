import openpyxl
import json
import codecs

wb = openpyxl.load_workbook('Calculatrice LSPD Spirit RP.xlsx', data_only=True)
sheet = wb.active

data = {}
current_cat = 'Autre'

for row in range(5, 300):
    cat_val = str(sheet.cell(row=row, column=3).value).strip()
    inf_val = str(sheet.cell(row=row, column=4).value).strip()
    amende_val = str(sheet.cell(row=row, column=5).value).strip()
    
    # Try looking for peines in column 8 or 9
    peine_val = str(sheet.cell(row=row, column=8).value).strip()
    if peine_val == 'None' or peine_val == '':
        peine_val = str(sheet.cell(row=row, column=9).value).strip()

    if cat_val in ['Contravention', 'Délit mineur', 'Délit Majeur', 'Crime', 'Crime Majeur', 'Délit', 'Crimes', 'Trafic']:
        current_cat = cat_val
        if current_cat not in data: data[current_cat] = []
        
    if inf_val != 'None' and inf_val != '' and inf_val != 'INFRACTION' and inf_val != '0':
        amende = 0
        peine = 0
        
        if amende_val != 'None' and amende_val != '':
            try:
                amende = int(float(amende_val))
            except:
                pass
                
        if peine_val != 'None' and peine_val != '':
            try:
                txt = peine_val.lower().replace('mois', '').replace(' ', '').strip()
                if txt.isdigit(): peine = int(txt)
            except:
                pass
                
        if current_cat not in data: data[current_cat] = []
        data[current_cat].append({'name': inf_val, 'amende': amende, 'peine': peine})

html = '<div class="calc-categories glass widget" style="max-height: 500px; overflow-y: auto;">\n'

for cat, items in data.items():
    if not items: continue
    
    html += f'    <div class="calc-group">\n'
    html += f'        <h4 class="calc-group-title">{cat}</h4>\n'
    
    seen = set()
    for item in items:
        name = item['name'].replace('"', '&quot;')
        if name in seen: continue
        seen.add(name)
        
        amende = item['amende']
        peine = item['peine']
        
        label_text = name
        if amende > 0 and peine > 0:
            label_text += f' (${amende}, {peine} mois)'
        elif amende > 0:
            label_text += f' (${amende})'
        elif peine > 0:
            label_text += f' ({peine} mois)'
            
        html += f'        <label class="calc-item"><input type="checkbox" value="{name}" data-amende="{amende}" data-prison="{peine}"> {label_text}</label>\n'
        
    html += f'    </div>\n'

html += '</div>'

with codecs.open('calc_html.txt', 'w', 'utf-8') as f:
    f.write(html)
